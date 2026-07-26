from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "students.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Students"
    sheet.append(["Roll No", "Name", "Marks"])
    wb.save(FILE_NAME)

# Load workbook
wb = load_workbook(FILE_NAME)
sheet = wb["Students"]

while True:

    print("\n===== STUDENT MANAGEMENT SYSTEM =====")
    print("1. Add Student")
    print("2. View Students")
    print("3. Search Student")
    print("4. Update Student")
    print("5. Delete Student")
    print("6. Exit")

    choice = input("Enter your choice: ")

    # Add Student
    if choice == "1":

        roll = input("Enter Roll Number: ")
        name = input("Enter Name: ")
        marks = input("Enter Marks: ")

        found = False

        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == roll:
                found = True
                break

        if found:
            print("Roll Number already exists.")
        else:
            sheet.append([roll, name, marks])
            wb.save(FILE_NAME)
            print("Student Added Successfully.")

    # View Students
    elif choice == "2":

        print("\nRoll No\tName\tMarks")
        print("-" * 30)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            print(f"{row[0]}\t{row[1]}\t{row[2]}")

    # Search Student
    elif choice == "3":

        roll = input("Enter Roll Number: ")

        found = False

        for row in sheet.iter_rows(min_row=2):
            if str(row[0].value) == roll:
                print("\nStudent Found")
                print("Roll No :", row[0].value)
                print("Name    :", row[1].value)
                print("Marks   :", row[2].value)
                found = True
                break

        if not found:
            print("Student Not Found.")

    # Update Student
    elif choice == "4":

        roll = input("Enter Roll Number: ")

        found = False

        for row in sheet.iter_rows(min_row=2):

            if str(row[0].value) == roll:

                row[1].value = input("Enter New Name: ")
                row[2].value = input("Enter New Marks: ")

                wb.save(FILE_NAME)

                print("Student Updated Successfully.")
                found = True
                break

        if not found:
            print("Student Not Found.")

    # Delete Student
    elif choice == "5":

        roll = input("Enter Roll Number: ")

        found = False

        for i in range(2, sheet.max_row + 1):

            if str(sheet.cell(i, 1).value) == roll:
                sheet.delete_rows(i)
                wb.save(FILE_NAME)
                print("Student Deleted Successfully.")
                found = True
                break

        if not found:
            print("Student Not Found.")

    # Exit
    elif choice == "6":
        wb.save(FILE_NAME)
        wb.close()
        print("Thank You!")
        break

    else:
        print("Invalid Choice.")